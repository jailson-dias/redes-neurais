from openpyxl import load_workbook
from random import shuffle

def ler_linhas(planilha, linhas, colunas):
	'''
	Ler as linhas de uma planilha.
	
	Args:
		planilha (openpyxl.worksheet.Worksheet): Planilha de onde as linhas serao lidas
		linhas (int): Numero de linhas a serem lidas
		colunas (int): Numero de colunas a serem lidas
	
	Yields:
		Generator de [float]: Um generator que a cada chamada produz uma lista com os valores presentes em uma linha da planilha.
	'''
	for row in planilha.iter_rows(min_row = 1, max_col = colunas, max_row = linhas):
		saida = []
		for cell in row:
			saida.append(cell.value)
		yield saida

#Separa a planilha em dois grupos, ja excluindo repeticoes em um mesmo grupo
def separa_grupos(planilha, grupo0, grupo1, linhas, colunas = 7):
	'''
	Separa as linhas de uma planilha em dois grupos com base no rotulo de cada linha (que e o ultimo valor), ja excluindo repeticoes em um mesmo grupo.
	
	Args:
		planilha (openpyxl.worksheet.Worksheet): Planilha de onde as linhas serao lidas
		grupo0 ([[float]]): Grupo onde todas as linhas com rotulo 0 ficarao
		grupo1 ([[float]]): Grupo onde todas as linhas com rotulo 1 ficarao
		linhas (int): Numero de linhas a serem lidas
		colunas (int): Numero de colunas a serem lidas
	'''
	print("Separando Grupos...")
	for linha in ler_linhas(planilha, linhas, colunas):
		rotulo = linha.pop(len(linha)-1)
		if (rotulo == 0) and (linha not in grupo0):
			grupo0.append(linha)
		elif (rotulo == 1) and (linha not in grupo1):
			grupo1.append(linha)
	#print("Linhas separadas: " + str(len(grupo0)+len(grupo1)))

def elimina_ruido(grupo0, grupo1):
	'''
	Elimina ruido nas amostras, removendo qualquer par de amostras que sejam exatamente iguais mas uma esteja presente em grupo0 e a outra esteja presente em grupo1.
	
	Args:
		grupo0 ([[float]]): Grupo onde os rotulos sao 0
		grupo1 ([[float]]): Grupo onde os rotulos sao 1
	'''
	print("Eliminando ruidos...")
	amostra0 = 0
	while amostra0 < len(grupo0):
		amostra1 = 0
		while amostra1 < len(grupo1):
			if grupo0[amostra0] == grupo1[amostra1]:
				grupo0.pop(amostra0)
				grupo1.pop(amostra1)
				break
			amostra1 = amostra1 + 1
		if amostra1 == len(grupo1):
			amostra0 = amostra0 + 1
		#print("Eliminando ruidos (" + str(amostra0) + "/" + str(len(grupo0)) + ")")

def insere_planilha(planilha, grupo, valor = ""):
	'''
	Insere as amostras do grupo na planilha.
	
	Args:
		planilha (openpyxl.worksheet.Worksheet): Planilha onde as amostras serao inseridas
		grupo ([[float]]): Grupo onde estao todas as amostras que serao inseridas
		valor (string): Rotulo das amostras
	'''
	print("Inserindo na planilha " + str(valor))
	for linha in range(len(grupo)):
		for col in range(len(grupo[linha])):
			planilha.cell(row = linha + 1, column = col + 1, value = grupo[linha][col])
		if (valor != ""):
			planilha.cell(row = linha + 1, column = len(grupo[linha]) + 1, value = valor)
		#print("Inseridas " + str(linha) + "/" + str(len(grupo)) + " linhas")

def dividir_em_grupos(grupoMaior = [], rotuloMaior = 0, grupoMenor = [], rotuloMenor = 1, numGrupos = 10):
	'''
	Divide os grupos em sets para serem divididos em treino, validacao e teste, repetindo amostras do grupoMenor para que cada grupo tenha uma proporcao equilibrada.
	
	Args:
		grupoMaior ([[float]]): O grupo com mais amostras
		rotuloMaior (int): O rotulo do grupoMaior
		grupoMenor ([[float]]): O grupo com menos amostras
		rotuloMenor (int): O rotulo do grupoMenor
		numGrupos: O numero de grupos a ser gerado
	
	Returns:
		[[[float]]]: Uma linha com <numGrupos> conjuntos contendo amostras de ambos os grupos
	'''
	grupos = []
	for i in range(numGrupos):
		grupos.append([])
	for i in range(len(grupoMaior)):
		grupoMaior[i].append(rotuloMaior)
		grupos[i % numGrupos].append(grupoMaior[i])
	
	print("Separacao do grupo maior")
	for grupo in grupos:
		print(len(grupo))
	
	for i in range(len(grupoMenor)):
		#O menor grupo tem insercoes repetidas para ficar do tamanho do maior
		grupoMenor[i].append(rotuloMenor)
		for j in range(len(grupoMaior) // len(grupoMenor)):
			grupos[i % numGrupos].append(grupoMenor[i])
	
	print("Separacao do grupo menor")
	for grupo in grupos:
		print(len(grupo))
	
	return grupos

xlsx = load_workbook("Entrada.xlsx")
grupoA = []
grupoB = []

separa_grupos(xlsx["Original"], grupoA, grupoB, 11183, 7)
print("Grupo 0: " + str(len(grupoA)))
print("Grupo 1: " + str(len(grupoB)))
elimina_ruido(grupoA, grupoB)
print("Grupo 0: " + str(len(grupoA)))
print("Grupo 1: " + str(len(grupoB)))
insere_planilha(xlsx.create_sheet(title = "Grupo 0"), grupoA, 0)
insere_planilha(xlsx.create_sheet(title = "Grupo 1"), grupoB, 1)

grupos = dividir_em_grupos(grupoMaior = grupoA, grupoMenor = grupoB, numGrupos = 10)
for i in range(len(grupos)):
	'''Embaralha e insere cada grupo em uma planilha propria'''
	shuffle(grupos[i])
	insere_planilha(xlsx.create_sheet(title = "Divisao " + str(i)), grupos[i])

xlsx.save("Saida.xlsx")