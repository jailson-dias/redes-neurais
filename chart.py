

from numpy import random
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import colors

from openpyxl import load_workbook
from math import ceil
from random import shuffle

def ler_linhas(planilha, linhas, colunas):
	'''
	Ler as linhas de uma planilha.
	
	Args:
		planilha (openpyxl.worksheet.Worksheet): Planilha de onde as linhas serao lidas
		linhas (int): Numero de linhas a serem lidas
		colunas (int): Numero de colunas a serem lidas
	
	Yields:
		Generator de [float]: Um generator que a cada chamada produz uma lista com os valores presentes em uma linha da planilha.
	'''
	for row in planilha.iter_rows(min_row = 1, max_col = colunas, max_row = linhas):
		saida = []
		for cell in row:
			saida.append(cell.value)
		yield saida

def separar_grupos(planilha, grupo0, grupo1):
    """
    ler a planilha com os dados e separa eles em grupo0 e grupo1
    """
    print ("Separando os dados em dois grupos")
    for linha in ler_linhas(planilha, 11183, 7):
        if linha[-1] == 0:
            grupo0.append(linha)
        else:
            grupo1.append(linha)


def eliminar_contradicao(grupo0, grupo1):

    """
    pega os dados dos grupos 0 e 1 e remove dos dois grupos os dados que são iguais
    mas estão nos dois grupos
    """

    print ("Eliminando contradições na base de dados")
    contradicao = 0
    l0 = len(grupo0) - 1
    while l0 >= 0:
        l1 = len(grupo1) - 1
        while (l1 >= 0):
            if (grupo0[l0][:-1] == grupo1[l1][:-1]):
                # print ("contradicao:",grupo0[l0], grupo1[l1])
                grupo0.pop(l0)
                grupo1.pop(l1)
                contradicao += 1
                break
            l1 -= 1
        l0 -= 1
    
    print("Contradições:", contradicao)

def dividir_em_grupos(grupo0, grupo1, quant_grupos = 10):
    """
    pega os dados do grupo0 e do grupo1 e separa ele em 10 grupos, sendo que em cada
    grupo vai ter a mesma quantidade de casos do grupo0 e do grupo1
    """

    print ("dividindo a base em " + str(quant_grupos) + " grupos com quantidade igual de casos de cada teste")
    def split_list(grupo, quant):
        grupos = []
        incremento = int(len(grupo)/quant)
        for g in range(0,len(grupo), incremento):
            grupos.append(grupo[g:g + incremento])
        return grupos[:quant]

    grupo0 = split_list(grupo0, quant_grupos)
    grupo1 = split_list(grupo1, quant_grupos)

    tam_grupo = max(len(grupo0[0]), len(grupo1[0]))

    for i in range(0,len(grupo0)):
        if len(grupo0[i]) < tam_grupo:
            grupo0[i] = grupo0[i] * ceil(tam_grupo/len(grupo0[i]))
            grupo0[i] = grupo0[i][:tam_grupo]
    
    for i in range(0,len(grupo1)):
        if len(grupo1[i]) < tam_grupo:
            grupo1[i] = grupo1[i] * ceil(tam_grupo/len(grupo1[i]))
            grupo1[i] = grupo1[i][:tam_grupo]

    grupos = []
    for i in range(0,quant_grupos):
        grupos.append(grupo0[i] + grupo1[i])
    
    return grupos

def insere_planilha(planilha, grupo):
	'''
	Insere as amostras do grupo na planilha.
	
	Args:
		planilha (openpyxl.worksheet.Worksheet): Planilha onde as amostras serao inseridas
		grupo ([[float]]): Grupo onde estao todas as amostras que serao inseridas
		valor (string): Rotulo das amostras
	'''
	for linha in range(len(grupo)):
		for col in range(len(grupo[linha])):
			planilha.cell(row = linha + 1, column = col + 1, value = grupo[linha][col])
       

def plotGraph(dados, title):
    print ("Desenhando o grafico da base de dados")
    cmap = colors.ListedColormap(['blue', 'red'])
    fig, axes = plt.subplots(nrows=1)
    axes.set_title(title, fontsize=14)
    # axes.axis([-0.5, 6.5, 0, 12000])
    # label = ["", "a","b", "c", "d", "e", "f", "g"]
    # frame1 = plt.gca()
    # frame1.axes.get_xaxis().set_ticklabels(label)
    dadosPlot = []
    for linha in dados:
        l = []
        for celula in linha[:-1]:
            if celula == 0 or celula == None:
                l.append(1)
            else:
                l.append(0)
        if linha[-1] != 0 and linha[-1] != None:
            l.append(0)
        else:
            l.append(1)
        dadosPlot.append(tuple(l))
    img = plt.imshow(dadosPlot, cmap=cmap, interpolation='nearest', aspect='auto')
    plt.show()

if __name__ == "__main__":


    xlsx = load_workbook("Entrada.xlsx")

    grupo0 = []
    grupo1 = []

    separar_grupos(xlsx["Original"], grupo0, grupo1)

    # print (len(grupo0), len(grupo1))
    eliminar_contradicao(grupo0, grupo1)
    # print (len(grupo0), len(grupo1))
    grupos = dividir_em_grupos(grupo0, grupo1)

    insere_planilha(xlsx.create_sheet(title = "Grupo 0"), grupo0)
    insere_planilha(xlsx.create_sheet(title = "Grupo 1"), grupo1)

    for i in range(len(grupos)):
        '''Embaralha e insere cada grupo em uma planilha propria'''
        shuffle(grupos[i])
        insere_planilha(xlsx.create_sheet(title = "Divisao " + str(i)), grupos[i])

    xlsx.save("Saida.xlsx")

    plotGraph(grupo0 + grupo1, "DADOS")